from openpyxl import Workbook, load_workbook

wb = load_workbook('files/CLIMA.xlsx')

ws = wb.active

columns = ['I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV']

for i in columns:
    for j in range(2,30):
        cell = ws[f'{i}{j}'].value
        if (cell == 'a. Siempre '):
            ws[f'{i}{j}'] = 5
        elif cell == 'b. Casi siempre ':
            ws[f'{i}{j}'] = 4
        elif cell == 'c. Algunas veces ':
            ws[f'{i}{j}'] = 3
        elif cell == 'd. Casi nunca ':
            ws[f'{i}{j}'] = 2
        elif cell == 'e. Nunca ':
            ws[f'{i}{j}'] = 1
        elif (cell == 'a. Totalmente de acuerdo '):
            ws[f'{i}{j}'] = 5
        elif (cell == 'b. Casi totalmente de acuerdo '):
            ws[f'{i}{j}'] = 4
        elif (cell == 'c. Indeciso '):
            ws[f'{i}{j}'] = 3
        elif (cell == 'd. Casi totalmente en desacuerdo '):
            ws[f'{i}{j}'] = 2
        elif (cell == 'e. Totalmente en desacuerdo '):
            ws[f'{i}{j}'] = 1
        elif (cell == 'a. Toda la información '):
            ws[f'{i}{j}'] = 5
        elif (cell == 'b. Algo de información '):
            ws[f'{i}{j}'] = 4
        elif (cell == 'c. Poca información '):
            ws[f'{i}{j}'] = 3
        elif (cell == 'd. Muy poca información '):
            ws[f'{i}{j}'] = 2
        elif (cell == 'e. Ninguna información '):
            ws[f'{i}{j}'] = 1

wb.save('files/CLIMA1.xlsx')