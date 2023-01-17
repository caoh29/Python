from openpyxl import Workbook, load_workbook

wb = load_workbook('files/CLIMA.xlsx')

ws = wb.active

columns = ['I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV']

for i in columns:
    for j in range(2,30):
        cell = ws[f'{i}{j}'].value
        string = str(cell)
        if (string.startswith('a')):
            ws[f'{i}{j}'] = 5
        elif (string.startswith('b')):
            ws[f'{i}{j}'] = 4
        elif (string.startswith('c')):
            ws[f'{i}{j}'] = 3
        elif (string.startswith('d')):
            ws[f'{i}{j}'] = 2
        elif (string.startswith('e')):
            ws[f'{i}{j}'] = 1

wb.save('files/CLIMA2.xlsx')